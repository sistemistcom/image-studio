import os
import re
import io
import json
import time
import zipfile
import mimetypes
import unicodedata
from datetime import datetime
from textwrap import dedent
from pathlib import Path
from urllib.parse import quote, urlparse

import requests
import boto3
from botocore.config import Config
from botocore.exceptions import ClientError, BotoCoreError

from PIL import Image, ImageOps
from openpyxl import load_workbook, Workbook

import streamlit as st


def get_config_value(name, default=""):
    """Render environment > Streamlit secrets > default."""
    value = os.getenv(name, "")
    if value:
        return str(value)
    try:
        value = st.secrets.get(name, default)
        return str(value) if value is not None else default
    except Exception:
        return default


# =========================================================
# SİSTEMİST IMAGE STUDIO WEB V8 STABLE
# =========================================================

os.environ["STREAMLIT_SERVER_ENABLE_CORS"] = "false"
os.environ["STREAMLIT_SERVER_ENABLE_XSRF_PROTECTION"] = "false"

st.set_page_config(
    page_title="Sistemist Image Studio",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="expanded"
)


# =========================================================
# SESSION STATE
# =========================================================

DEFAULTS = {
    "current_page": "Dashboard",
    "history": [],
    # R2 bilgileri Render Environment Variables veya Streamlit secrets'tan
    # otomatik okunabilir. Panelden girilen değerler sadece aktif oturumda tutulur.
    "r2_endpoint": get_config_value("R2_ENDPOINT", ""),
    "r2_access_key": get_config_value("R2_ACCESS_KEY_ID", ""),
    "r2_secret_key": get_config_value("R2_SECRET_ACCESS_KEY", ""),
    "r2_bucket": get_config_value("R2_BUCKET", "sistemist-image-studio"),
    "r2_public_url": get_config_value("R2_PUBLIC_URL", ""),
    "r2_region": get_config_value("R2_REGION", "auto"),
    "last_processed": 0,
    "last_success": 0,
    "active_package": "PRO",
}

for key, value in DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = value


# =========================================================
# GLOBAL CSS
# =========================================================

st.markdown(dedent("""
<style>

/* ---------------------------------------------------------
   GOOGLE FONT
--------------------------------------------------------- */

@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=Space+Grotesk:wght@500;600;700&display=swap');


/* ---------------------------------------------------------
   ROOT
--------------------------------------------------------- */

:root {
    --bg: #0b1119;
    --sidebar: #111923;
    --panel: #151f2b;
    --panel2: #192432;
    --border: #2a394b;
    --text: #f4f7fb;
    --muted: #8b9aab;
    --orange: #ff6a00;
    --orange2: #ff8a2a;
    --blue: #4da3ff;
    --green: #35d49a;
    --danger: #ff5c6c;
}


/* ---------------------------------------------------------
   APP
--------------------------------------------------------- */

.stApp {
    background:
        radial-gradient(circle at 85% 0%, rgba(255,106,0,0.08), transparent 28%),
        radial-gradient(circle at 20% 100%, rgba(48,110,255,0.05), transparent 30%),
        #0b1119 !important;
    color: var(--text) !important;
    font-family: 'Inter', sans-serif !important;
}

.main .block-container {
    padding-top: 2rem !important;
    padding-bottom: 4rem !important;
    max-width: 1500px !important;
}


/* ---------------------------------------------------------
   HIDE DEFAULT STREAMLIT ELEMENTS
--------------------------------------------------------- */

#MainMenu,
footer,
[data-testid="stHeader"] {
    display: none !important;
}


/* ---------------------------------------------------------
   SIDEBAR
--------------------------------------------------------- */

[data-testid="stSidebar"] {
    width: 290px !important;
    min-width: 290px !important;
    background:
        linear-gradient(180deg, #141d28 0%, #101720 100%) !important;
    border-right: 1px solid #263545 !important;
}

[data-testid="stSidebar"] > div:first-child {
    padding: 0 !important;
}

.sidebar-wrap {
    padding: 32px 22px 25px 22px;
}

.sidebar-brand {
    padding-bottom: 28px;
    border-bottom: 1px solid #273545;
    margin-bottom: 28px;
}

.brand-row {
    display: flex;
    align-items: center;
    gap: 13px;
}

.brand-symbol {
    width: 46px;
    height: 46px;
    position: relative;
    flex-shrink: 0;
}

.brand-symbol::before {
    content: "";
    position: absolute;
    width: 33px;
    height: 33px;
    top: 1px;
    left: 4px;
    border-radius: 10px 10px 4px 10px;
    background: linear-gradient(135deg, #ff8a28, #ff5500);
    transform: rotate(45deg);
}

.brand-symbol::after {
    content: "";
    position: absolute;
    width: 25px;
    height: 25px;
    top: 15px;
    left: 16px;
    border-radius: 5px 10px 10px 4px;
    background: #eef2f7;
    transform: rotate(45deg);
}

.brand-name {
    color: #ffffff;
    font-size: 25px;
    font-weight: 800;
    letter-spacing: 4px;
    line-height: 1;
}

.brand-name span {
    color: var(--orange);
}

.brand-version {
    color: #718399;
    font-size: 9px;
    font-weight: 700;
    letter-spacing: 2px;
    margin-top: 10px;
    margin-left: 59px;
}

.nav-label {
    color: #718399;
    font-size: 9px;
    font-weight: 800;
    letter-spacing: 2px;
    margin: 22px 8px 9px 8px;
    text-transform: uppercase;
}


/* ---------------------------------------------------------
   SIDEBAR BUTTONS
--------------------------------------------------------- */

[data-testid="stSidebar"] .stButton {
    margin-bottom: 5px;
}

[data-testid="stSidebar"] .stButton > button {
    width: 100%;
    min-height: 46px;
    background: transparent !important;
    color: #aebdce !important;
    border: 1px solid transparent !important;
    border-radius: 10px !important;
    box-shadow: none !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    text-align: left !important;
    justify-content: flex-start !important;
    padding-left: 14px !important;
    transition: all .2s ease !important;
}

[data-testid="stSidebar"] .stButton > button:hover {
    background: #1b2735 !important;
    border-color: #2c3d50 !important;
    color: #ffffff !important;
    transform: translateX(3px);
}

.sidebar-bottom {
    margin-top: 28px;
    padding-top: 22px;
    border-top: 1px solid #273545;
}

.sidebar-status {
    display: flex;
    align-items: center;
    gap: 9px;
    background: #172230;
    border: 1px solid #293b4e;
    border-radius: 12px;
    padding: 13px;
}

.status-dot {
    width: 8px;
    height: 8px;
    background: #32d583;
    border-radius: 50%;
    box-shadow: 0 0 12px rgba(50,213,131,.8);
}

.status-text {
    color: #d5deea;
    font-size: 12px;
    font-weight: 600;
}

.status-sub {
    color: #718399;
    font-size: 10px;
    margin-top: 3px;
}


/* ---------------------------------------------------------
   GENERAL TEXT
--------------------------------------------------------- */

h1, h2, h3, h4 {
    font-family: 'Space Grotesk', 'Inter', sans-serif !important;
}

h1 {
    color: #f5f7fa !important;
}

p, span, label {
    font-family: 'Inter', sans-serif !important;
}


/* ---------------------------------------------------------
   HERO
--------------------------------------------------------- */

.hero {
    position: relative;
    overflow: hidden;
    padding: 34px 36px;
    border-radius: 20px;
    border: 1px solid #27384a;
    background:
        radial-gradient(circle at 85% 10%, rgba(255,106,0,.16), transparent 30%),
        linear-gradient(135deg, #151f2b, #101721);
    margin-bottom: 24px;
}

.hero::after {
    content: "";
    position: absolute;
    width: 280px;
    height: 280px;
    right: -120px;
    top: -170px;
    border: 1px solid rgba(255,106,0,.12);
    border-radius: 50%;
    box-shadow:
        0 0 0 50px rgba(255,106,0,.025),
        0 0 0 100px rgba(255,106,0,.015);
}

.system-read {
    color: var(--orange);
    font-size: 10px;
    letter-spacing: 3px;
    font-weight: 800;
    margin-bottom: 12px;
    position: relative;
    z-index: 2;
}

.hero-title {
    color: #f5f7fb;
    font-size: 35px;
    font-weight: 700;
    letter-spacing: -.8px;
    margin: 0;
    position: relative;
    z-index: 2;
}

.hero-title span {
    color: var(--orange);
}

.hero-subtitle {
    max-width: 720px;
    color: #8fa0b3;
    font-size: 14px;
    line-height: 1.8;
    margin-top: 13px;
    position: relative;
    z-index: 2;
}


/* ---------------------------------------------------------
   STAT CARDS
--------------------------------------------------------- */

.stat-card {
    position: relative;
    overflow: hidden;
    min-height: 155px;
    background: linear-gradient(135deg, #17212d, #131c27);
    border: 1px solid #2a3a4c;
    border-radius: 17px;
    padding: 21px;
    transition: all .2s ease;
}

.stat-card:hover {
    border-color: #40556b;
    transform: translateY(-2px);
}

.stat-card.orange {
    border-left: 4px solid var(--orange);
}

.stat-icon {
    color: var(--orange);
    font-size: 20px;
    margin-bottom: 22px;
}

.stat-label {
    color: #718399;
    font-size: 10px;
    font-weight: 800;
    letter-spacing: 1px;
    text-transform: uppercase;
}

.stat-value {
    color: #f2f6fa;
    font-family: 'Space Grotesk', sans-serif;
    font-size: 28px;
    font-weight: 700;
    margin-top: 8px;
}

.stat-sub {
    color: #66798d;
    font-size: 11px;
    margin-top: 9px;
}


/* ---------------------------------------------------------
   PANELS
--------------------------------------------------------- */

.panel {
    background:
        radial-gradient(circle at 100% 0%, rgba(255,106,0,.05), transparent 28%),
        #151f2a;
    border: 1px solid #2a3a4c;
    border-radius: 20px;
    padding: 27px;
    margin-top: 20px;
}

.panel-title {
    color: #f0f4f8;
    font-size: 21px;
    font-weight: 700;
    margin-bottom: 8px;
}

.panel-subtitle {
    color: #8294a8;
    font-size: 13px;
    line-height: 1.7;
    margin-bottom: 24px;
}


/* ---------------------------------------------------------
   ENGINE CARDS
--------------------------------------------------------- */

.engine-card {
    position: relative;
    overflow: hidden;
    height: 100%;
    min-height: 260px;
    padding: 28px;
    background:
        radial-gradient(circle at 100% 0%, rgba(255,106,0,.09), transparent 28%),
        #17212c;
    border: 1px solid #2a3b4d;
    border-radius: 20px;
}

.engine-card::after {
    content: "";
    position: absolute;
    width: 150px;
    height: 150px;
    right: -70px;
    top: -70px;
    border-radius: 50%;
    background: rgba(255,106,0,.035);
}

.engine-icon {
    width: 56px;
    height: 56px;
    display: flex;
    align-items: center;
    justify-content: center;
    background: rgba(255,106,0,.08);
    border: 1px solid rgba(255,106,0,.24);
    color: var(--orange);
    border-radius: 15px;
    font-size: 23px;
    margin-bottom: 22px;
}

.engine-title {
    color: #f2f5f8;
    font-size: 21px;
    font-weight: 700;
}

.engine-text {
    color: #899aac;
    font-size: 13px;
    line-height: 1.8;
    margin-top: 12px;
}


/* ---------------------------------------------------------
   MAIN BUTTONS
--------------------------------------------------------- */

.stButton > button {
    min-height: 46px !important;
    border-radius: 11px !important;
    border: 1px solid #ff6a00 !important;
    background: linear-gradient(135deg, #ff7a18, #ff5b00) !important;
    color: #ffffff !important;
    font-size: 13px !important;
    font-weight: 700 !important;
    box-shadow: 0 8px 22px rgba(255,106,0,.16) !important;
    transition: all .2s ease !important;
}

.stButton > button:hover {
    border-color: #ff8c3b !important;
    background: linear-gradient(135deg, #ff8b30, #ff630b) !important;
    transform: translateY(-1px);
    box-shadow: 0 12px 28px rgba(255,106,0,.22) !important;
}


/* ---------------------------------------------------------
   INPUTS
--------------------------------------------------------- */

.stTextInput input,
.stSelectbox > div > div,
.stNumberInput input,
.stTextArea textarea {
    background: #101821 !important;
    color: #edf3f8 !important;
    border: 1px solid #304154 !important;
    border-radius: 10px !important;
}

.stTextInput input:focus,
.stNumberInput input:focus,
.stTextArea textarea:focus {
    border-color: var(--orange) !important;
    box-shadow: 0 0 0 1px var(--orange) !important;
}

.stTextInput label,
.stSelectbox label,
.stNumberInput label,
.stTextArea label {
    color: #9baabd !important;
    font-size: 12px !important;
    font-weight: 600 !important;
}


/* ---------------------------------------------------------
   FILE UPLOADER
--------------------------------------------------------- */

[data-testid="stFileUploader"] {
    background: #111a24;
    border: 1px dashed #3a4d61;
    border-radius: 16px;
    padding: 15px;
}

[data-testid="stFileUploader"]:hover {
    border-color: var(--orange);
}


/* ---------------------------------------------------------
   EXPANDER
--------------------------------------------------------- */

[data-testid="stExpander"] {
    background: #121b25;
    border: 1px solid #2b3b4d;
    border-radius: 14px;
}

[data-testid="stExpander"] summary {
    color: #e5edf5 !important;
}


/* ---------------------------------------------------------
   ALERTS
--------------------------------------------------------- */

.stSuccess,
.stInfo,
.stWarning,
.stError {
    border-radius: 12px !important;
}


/* ---------------------------------------------------------
   TABLE
--------------------------------------------------------- */

[data-testid="stDataFrame"] {
    border: 1px solid #2b3c4f;
    border-radius: 14px;
    overflow: hidden;
}


/* ---------------------------------------------------------
   DIVIDER
--------------------------------------------------------- */

hr {
    border-color: #263546 !important;
}


/* ---------------------------------------------------------
   FOOTER
--------------------------------------------------------- */

.app-footer {
    margin-top: 45px;
    padding-top: 20px;
    border-top: 1px solid #243345;
    color: #5e7185;
    font-size: 10px;
    letter-spacing: 1px;
    text-align: center;
}


/* V8 DASHBOARD POLISH */
[data-testid="stSidebar"]{width:285px!important;min-width:285px!important;background:linear-gradient(180deg,#121c29 0%,#0d1520 100%)!important;}
.sidebar-wrap{padding:28px 26px 0;}.sidebar-brand{padding-bottom:23px;border-bottom:1px solid #263647}.brand-symbol{width:48px;height:48px;border-radius:14px;background:linear-gradient(135deg,#ff8a2a,#ff5900);display:flex;align-items:center;justify-content:center;box-shadow:0 12px 28px rgba(255,106,0,.2)}.brand-symbol:before{content:"S";color:#fff;font-size:24px;font-weight:800;font-family:'Space Grotesk',sans-serif}.brand-symbol:after{display:none!important}.brand-name{font-size:24px;letter-spacing:3px}.brand-name span{color:#fff}.brand-version{color:#ff9b4c;margin:8px 0 0 61px;font-size:10px;letter-spacing:3px}.nav-label{font-size:10px;letter-spacing:3px;margin:27px 12px 10px}
[data-testid="stSidebar"] .stButton{margin:0 0 5px}[data-testid="stSidebar"] .stButton>button{min-height:48px;background:transparent!important;color:#c3ceda!important;border:1px solid transparent!important;border-radius:11px!important;box-shadow:none!important;justify-content:flex-start!important;text-align:left!important}[data-testid="stSidebar"] .stButton>button:hover{background:#182534!important;border-color:#2a3d50!important;color:#fff!important;transform:translateX(2px)!important}[data-testid="stSidebar"] .stButton>button[kind="primary"]{color:#fff!important;border-color:#a84c0b!important;background:linear-gradient(90deg,#8d3e0a 0%,#c45a0b 58%,#ff6a00 100%)!important;box-shadow:0 8px 24px rgba(255,106,0,.18)!important}
.sidebar-bottom{margin:24px 26px!important;padding:18px!important;border:1px solid #314254!important;border-radius:16px!important;background:linear-gradient(145deg,#192432,#121b26)!important}.sidebar-status{display:block!important;padding:0!important;background:transparent!important;border:0!important}.status-dot{display:inline-block;margin-right:7px}.status-text{display:inline}.status-sub{margin-top:9px}
.dashboard-heading{margin:0 0 26px}.page-kicker{color:#ff9b4c;font-size:10px;font-weight:800;letter-spacing:3px;text-transform:uppercase;margin-bottom:10px}.page-title{color:#f7f9fc;font-family:'Space Grotesk',sans-serif;font-size:34px;font-weight:700;letter-spacing:-.8px;margin:0}.page-subtitle{color:#98a8ba;font-size:14px;line-height:1.75;margin-top:8px}.dash-card{min-height:255px;padding:28px 29px;border-radius:20px;background:linear-gradient(145deg,#172230,#111a25);border:1px solid #2b3d50;position:relative;overflow:hidden}.dash-card:after{content:"";position:absolute;right:-45px;top:-45px;width:150px;height:150px;border-radius:50%;background:rgba(255,106,0,.045)}.dash-icon{width:58px;height:58px;border-radius:16px;display:flex;align-items:center;justify-content:center;background:rgba(255,106,0,.09);border:1px solid rgba(255,138,42,.28);font-size:25px;margin-bottom:22px;color:#ff8a2a}.dash-title{font-size:21px;font-weight:700;color:#f4f7fb;margin-bottom:10px}.dash-text{font-size:13px;line-height:1.8;color:#95a5b7;max-width:470px}.dashboard-action .stButton>button{width:100%!important;margin-top:0!important}
[data-testid="stFileUploader"]{background:#101923!important;border:1px dashed #4a6077!important;border-radius:18px!important;padding:16px!important}[data-testid="stFileUploader"] section{background:transparent!important;color:#dce6f0!important}[data-testid="stFileUploader"] button{background:linear-gradient(135deg,#ff7a18,#ff5b00)!important;color:#fff!important;border:1px solid #ff7a18!important;border-radius:10px!important;font-weight:700!important}[data-testid="stFileUploader"] small,[data-testid="stFileUploader"] span,[data-testid="stFileUploader"] div{color:#b8c5d3!important}

</style>
"""), unsafe_allow_html=True)


# =========================================================
# HELPERS
# =========================================================

def clean_filename(value):
    value = str(value or "urun").strip()

    replacements = {
        "ç": "c", "Ç": "C",
        "ğ": "g", "Ğ": "G",
        "ı": "i", "İ": "I",
        "ö": "o", "Ö": "O",
        "ş": "s", "Ş": "S",
        "ü": "u", "Ü": "U"
    }

    for old, new in replacements.items():
        value = value.replace(old, new)

    value = unicodedata.normalize("NFKD", value)
    value = "".join(
        c for c in value
        if not unicodedata.combining(c)
    )

    value = re.sub(r'[<>:"/\\\\|?*]', "-", value)
    value = re.sub(r"\s+", "-", value)
    value = re.sub(r"-+", "-", value)

    value = value.strip(" .-_")

    return value[:120] or "urun"


def is_url(value):
    if not isinstance(value, str):
        return False

    return value.strip().lower().startswith(
        ("http://", "https://")
    )


def format_size(size_bytes):
    if not size_bytes:
        return "0 B"

    units = ["B", "KB", "MB", "GB"]

    size = float(size_bytes)

    for unit in units:
        if size < 1024:
            return f"{size:.1f} {unit}"
        size /= 1024

    return f"{size:.1f} TB"


def get_file_extension_from_url(url):
    try:
        path = urlparse(url).path
        extension = Path(path).suffix.lower()

        if extension in [
            ".jpg", ".jpeg", ".png",
            ".webp", ".gif", ".bmp",
            ".tif", ".tiff", ".avif"
        ]:
            return extension

    except Exception:
        pass

    return ".jpg"


def add_history(operation, status, detail, count=0):
    record = {
        "Tarih": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
        "İşlem": operation,
        "Durum": status,
        "Detay": detail,
        "Dosya": count
    }

    st.session_state.history.insert(0, record)

    st.session_state.history = (
        st.session_state.history[:100]
    )


def read_image_excel(file_bytes):
    workbook = load_workbook(
        io.BytesIO(file_bytes),
        read_only=True,
        data_only=True
    )

    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        first_row = next(rows)
    except StopIteration:
        workbook.close()
        raise RuntimeError("Excel dosyası boş.")

    headers = [
        str(value).strip()
        if value is not None
        else ""
        for value in first_row
    ]

    data = []

    for row in rows:
        row_data = {}

        for index, header in enumerate(headers):
            if not header:
                continue

            row_data[header] = (
                row[index]
                if index < len(row)
                else None
            )

        data.append(row_data)

    workbook.close()

    image_columns = []

    for header in headers:

        normalized = (
            header.upper()
            .replace("İ", "I")
            .replace("Ş", "S")
            .replace("Ü", "U")
        )

        if (
            normalized.startswith("RESIM")
            or normalized.startswith("GÖRSEL")
            or normalized.startswith("GORSEL")
            or normalized.startswith("IMAGE")
        ):
            image_columns.append(header)

    def image_sort_key(column):
        match = re.search(r"\d+", column)

        if match:
            return int(match.group())

        return 9999

    image_columns.sort(key=image_sort_key)

    return headers, data, image_columns


def flatten_to_rgb(image, background="white"):
    if image.mode in ("RGBA", "LA"):
        rgba = image.convert("RGBA")

        background_image = Image.new(
            "RGB",
            rgba.size,
            background
        )

        background_image.paste(
            rgba,
            mask=rgba.getchannel("A")
        )

        return background_image

    if image.mode not in ("RGB", "L"):
        return image.convert("RGB")

    return image.convert("RGB")


def prepare_image(image, target_size, fit_mode):
    try:
        image.seek(0)
    except Exception:
        pass

    try:
        image = ImageOps.exif_transpose(image)
    except Exception:
        pass

    if target_size is None:
        return image.copy()

    target_width, target_height = target_size

    if fit_mode == "Kırp":

        processed = ImageOps.fit(
            image,
            (target_width, target_height),
            method=Image.Resampling.LANCZOS,
            centering=(0.5, 0.5)
        )

        return processed

    processed = image.copy()

    processed.thumbnail(
        (target_width, target_height),
        Image.Resampling.LANCZOS
    )

    processed = flatten_to_rgb(processed)

    canvas = Image.new(
        "RGB",
        (target_width, target_height),
        "white"
    )

    x = (target_width - processed.width) // 2
    y = (target_height - processed.height) // 2

    canvas.paste(processed, (x, y))

    return canvas


def get_target_size(size_mode):

    sizes = {
        "1200 × 1200 px": (1200, 1200),
        "1200 × 1800 px": (1200, 1800),
        "1000 × 1000 px": (1000, 1000),
        "800 × 800 px": (800, 800),
        "1920 × 1920 px": (1920, 1920),
        "Orijinal Boyut": None
    }

    return sizes.get(size_mode)


def save_image_to_buffer(image, output_format, quality=90):

    buffer = io.BytesIO()

    format_map = {
        "JPG": ("JPEG", ".jpg"),
        "PNG": ("PNG", ".png"),
        "WEBP": ("WEBP", ".webp"),
    }

    if output_format not in format_map:
        output_format = "JPG"

    pil_format, extension = format_map[output_format]

    if pil_format == "JPEG":
        image = flatten_to_rgb(image)

        image.save(
            buffer,
            format="JPEG",
            quality=int(quality),
            optimize=True
        )

    elif pil_format == "PNG":

        if image.mode not in ("RGB", "RGBA"):
            image = image.convert("RGBA")

        image.save(
            buffer,
            format="PNG",
            optimize=True
        )

    elif pil_format == "WEBP":

        image = flatten_to_rgb(image)

        image.save(
            buffer,
            format="WEBP",
            quality=int(quality),
            method=6
        )

    buffer.seek(0)

    return buffer.getvalue(), extension


def get_r2_client():

    endpoint = str(st.session_state.r2_endpoint).strip()
    access_key = str(st.session_state.r2_access_key).strip()
    secret_key = str(st.session_state.r2_secret_key).strip()

    if not endpoint:
        raise RuntimeError("R2 Endpoint girilmemiş.")

    if not access_key:
        raise RuntimeError("Access Key ID girilmemiş.")

    if not secret_key:
        raise RuntimeError("Secret Access Key girilmemiş.")

    return boto3.client(
        "s3",
        endpoint_url=endpoint.rstrip("/"),
        aws_access_key_id=access_key,
        aws_secret_access_key=secret_key,
        region_name=st.session_state.r2_region or "auto",
        config=Config(
            signature_version="s3v4",
            retries={
                "max_attempts": 3,
                "mode": "standard"
            }
        )
    )


def r2_is_configured():

    return all([
        str(st.session_state.r2_endpoint).strip(),
        str(st.session_state.r2_access_key).strip(),
        str(st.session_state.r2_secret_key).strip(),
        str(st.session_state.r2_bucket).strip(),
        str(st.session_state.r2_public_url).strip()
    ])


def build_public_url(object_key):

    public_url = (
        st.session_state.r2_public_url
        .rstrip("/")
    )

    quoted_key = quote(
        object_key,
        safe="/"
    )

    return f"{public_url}/{quoted_key}"


def page_header(title, subtitle, eyebrow="SİSTEMİST IMAGE STUDIO"):

    st.markdown(
        dedent(f"""
        <div class="hero">
            <div class="system-read">{eyebrow}</div>
            <h1 class="hero-title">{title}</h1>
            <div class="hero-subtitle">{subtitle}</div>
        </div>
        """),
        unsafe_allow_html=True
    )


def app_footer():

    st.markdown(
        dedent("""
        <div class="app-footer">
            © 2026 SİSTEMİST IMAGE STUDIO • PROFESSIONAL SAAS PLATFORM
        </div>
        """),
        unsafe_allow_html=True
    )


def go_to(page):
    st.session_state.current_page = page


def nav_button(label, page, key):
    active = st.session_state.current_page == page
    if st.button(label, key=key, type="primary" if active else "secondary"):
        st.session_state.current_page = page
        st.rerun()


# =========================================================
# SIDEBAR
# =========================================================

with st.sidebar:
    st.markdown(dedent("""
        <div class="sidebar-wrap">
            <div class="sidebar-brand">
                <div class="brand-row">
                    <div class="brand-symbol"></div>
                    <div>
                        <div class="brand-name">SİSTEMİST</div>
                        <div class="brand-version">IMAGE STUDIO</div>
                    </div>
                </div>
            </div>
        </div>
    """), unsafe_allow_html=True)

    st.markdown('<div class="nav-label">Çalışma Alanı</div>', unsafe_allow_html=True)
    nav_button("⌂  Dashboard", "Dashboard", "nav_dashboard")
    nav_button("↙  URL → Görsel", "URL → Görsel", "nav_url_image")
    nav_button("↗  Görsel → URL", "Görsel → URL", "nav_image_url")
    nav_button("◇  Toplu Dönüştürme", "Toplu Dönüştürme", "nav_batch")
    nav_button("◷  İşlem Geçmişi", "İşlem Geçmişi", "nav_history")

    st.markdown('<div class="nav-label">Sistem</div>', unsafe_allow_html=True)
    nav_button("☁  Cloud Dosyaları", "Cloud Dosyaları", "nav_cloud_files")
    nav_button("⚙  Cloud R2 Ayarları", "Cloud R2 Ayarları", "nav_r2")
    nav_button("◉  Genel Ayarlar", "Genel Ayarlar", "nav_settings")

    st.markdown('<div class="nav-label">Destek</div>', unsafe_allow_html=True)
    nav_button("?  Yardım Merkezi", "Yardım Merkezi", "nav_help")
    nav_button("◆  Paket & Lisans", "Paket & Lisans", "nav_package")

    st.markdown(dedent("""
        <div class="sidebar-bottom">
            <div class="sidebar-status">
                <div><span class="status-dot"></span><span class="status-text">Sistem Aktif</span></div>
                <div class="status-sub">Image Studio hizmete hazır</div>
            </div>
        </div>
    """), unsafe_allow_html=True)


# =========================================================
# DASHBOARD
# =========================================================

if st.session_state.current_page == "Dashboard":

    total_history = len(st.session_state.history)
    total_files = sum(item.get("Dosya", 0) for item in st.session_state.history)
    success_files = sum(item.get("Dosya", 0) for item in st.session_state.history if item.get("Durum") == "Başarılı")
    success_rate = round((success_files / total_files) * 100, 1) if total_files else 0
    r2_ready = r2_is_configured()

    st.markdown("""
        <div class="dashboard-heading">
            <div class="page-kicker">SİSTEMİST IMAGE STUDIO</div>
            <div class="page-title">Image Studio</div>
            <div class="page-subtitle">E-ticaret görsellerinizi indirin, işleyin ve buluta yükleyin. Tüm görsel operasyonlarınız tek panelde.</div>
        </div>
    """, unsafe_allow_html=True)

    st.markdown("""
        <div class="hero">
            <div class="system-read">SİSTEMİST IMAGE STUDIO</div>
            <h1 class="hero-title">Görsel operasyonlarınızı tek merkezden yönetin.</h1>
            <div class="hero-subtitle">Excel'deki ürün görsellerini toplu indirin, yeniden boyutlandırın, ZIP oluşturun veya görsellerinizi Cloudflare R2'ye yükleyerek doğrudan kullanılabilir URL'ler oluşturun.</div>
        </div>
    """, unsafe_allow_html=True)

    engine1, engine2 = st.columns(2, gap="large")
    with engine1:
        st.markdown("""<div class="dash-card"><div class="dash-icon">↓</div><div class="dash-title">URL → Görsel Motoru</div><div class="dash-text">Excel dosyanızdaki ürün görsel bağlantılarını otomatik olarak indirin. JPG, PNG veya WEBP dönüşümü yapın. İstediğiniz ölçüde görselleri işleyip tek ZIP dosyasında alın.</div></div>""", unsafe_allow_html=True)
        st.markdown('<div class="dashboard-action">', unsafe_allow_html=True)
        if st.button("URL → GÖRSEL MOTORUNU AÇ", key="open_url_engine"):
            go_to("URL → Görsel"); st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    with engine2:
        st.markdown("""<div class="dash-card"><div class="dash-icon">☁</div><div class="dash-title">Görsel → URL Motoru</div><div class="dash-text">Yerel görsellerinizi doğrudan Cloudflare R2 depolamaya yükleyin. İşlem tamamlandığında tüm görsel URL'lerini içeren hazır Excel raporunu tek tıklamayla indirin.</div></div>""", unsafe_allow_html=True)
        st.markdown('<div class="dashboard-action">', unsafe_allow_html=True)
        if st.button("GÖRSEL → URL MOTORUNU AÇ", key="open_image_engine"):
            go_to("Görsel → URL"); st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
    c1,c2,c3,c4 = st.columns(4, gap="medium")
    stats=[("⌁","Toplam Dönüştürme",f"{total_files:,}".replace(",","."),"Tüm zamanlar"),("✓","Başarılı İşlem",f"{success_files:,}".replace(",","."),f"%{success_rate} başarı oranı"),("◷","Kayıtlı Operasyon",str(total_history),"İşlem geçmişi"),("☁","Cloud R2","HAZIR" if r2_ready else "AYARLA","Depolama bağlantısı")]
    for column,(icon,label,value,sub) in zip([c1,c2,c3,c4],stats):
        with column:
            st.markdown(f'<div class="stat-card"><div class="stat-icon">{icon}</div><div class="stat-label">{label}</div><div class="stat-value">{value}</div><div class="stat-sub">{sub}</div></div>', unsafe_allow_html=True)

    st.markdown('<div class="panel"><div class="panel-title">Son İşlemler</div><div class="panel-subtitle">Sistem üzerinde gerçekleştirilen en son operasyonlar.</div></div>', unsafe_allow_html=True)
    if st.session_state.history:
        st.dataframe(st.session_state.history[:8], use_container_width=True, hide_index=True)
        if st.button("TÜM İŞLEM GEÇMİŞİNİ GÖR", key="dashboard_history"):
            go_to("İşlem Geçmişi"); st.rerun()
    else:
        st.info("Henüz işlem geçmişi bulunmuyor. URL → Görsel veya Görsel → URL aracını kullanarak başlayabilirsiniz.")
    app_footer()


# =========================================================
# URL -> IMAGE
# =========================================================

elif st.session_state.current_page == "URL → Görsel":

    page_header(
        "<span>URL → Görsel</span> İşleme Merkezi",
        "Excel dosyanızdaki görsel URL'lerini otomatik olarak indirin, dönüştürün ve tek bir ZIP dosyasında toplayın.",
        "SİSTEMİST IMAGE ENGINE"
    )

    uploaded_excel = st.file_uploader(
        "Excel dosyasını yükleyin",
        type=["xlsx"],
        key="url_excel_uploader"
    )

    if uploaded_excel:

        try:

            file_bytes = uploaded_excel.getvalue()

            headers, excel_data, image_columns = read_image_excel(
                file_bytes
            )

            st.success(
                f"Excel başarıyla analiz edildi. {len(excel_data)} ürün satırı ve {len(image_columns)} görsel kolonu bulundu."
            )

            if not image_columns:

                st.warning(
                    "RESIM1, RESIM2, GÖRSEL1, IMAGE1 benzeri görsel URL kolonları bulunamadı."
                )

            else:

                st.markdown(
                    dedent("""
                    <div class="panel">
                        <div class="panel-title">İşlem Ayarları</div>
                        <div class="panel-subtitle">
                            İndirilecek görsellerin formatını ve ölçülerini seçin.
                        </div>
                    </div>
                    """),
                    unsafe_allow_html=True
                )

                col1, col2, col3 = st.columns(3)

                usable_headers = [
                    header
                    for header in headers
                    if header and header not in image_columns
                ]

                with col1:

                    name_col = st.selectbox(
                        "Dosya adı sütunu",
                        usable_headers
                        if usable_headers
                        else headers
                    )

                with col2:

                    output_format = st.selectbox(
                        "Dönüşüm formatı",
                        ["JPG", "PNG", "WEBP"]
                    )

                with col3:

                    size_mode = st.selectbox(
                        "Görsel boyutu",
                        [
                            "1200 × 1200 px",
                            "1200 × 1800 px",
                            "1000 × 1000 px",
                            "800 × 800 px",
                            "1920 × 1920 px",
                            "Orijinal Boyut"
                        ]
                    )

                col4, col5 = st.columns(2)

                with col4:

                    fit_mode_label = st.selectbox(
                        "Yerleşim modu",
                        [
                            "Sığdır",
                            "Kırp"
                        ]
                    )

                with col5:

                    quality = st.slider(
                        "Görsel kalitesi",
                        min_value=60,
                        max_value=100,
                        value=90
                    )

                if st.button(
                    "GÖRSELLERİ İŞLE VE ZIP OLUŞTUR",
                    key="process_url_images"
                ):

                    tasks = []

                    for row_number, row in enumerate(
                        excel_data,
                        start=2
                    ):

                        base_name = clean_filename(
                            row.get(name_col)
                            or f"urun-{row_number}"
                        )

                        image_number = 0

                        for image_column in image_columns:

                            value = row.get(image_column)

                            if is_url(value):

                                image_number += 1

                                tasks.append({
                                    "url": value.strip(),
                                    "base": base_name,
                                    "image_number": image_number
                                })

                    if not tasks:

                        st.warning(
                            "Excel içerisinde geçerli görsel URL'si bulunamadı."
                        )

                    else:

                        target_size = get_target_size(size_mode)

                        zip_buffer = io.BytesIO()

                        success_count = 0
                        failed_count = 0
                        errors = []

                        progress = st.progress(0)
                        status = st.empty()

                        session = requests.Session()

                        with zipfile.ZipFile(
                            zip_buffer,
                            "w",
                            zipfile.ZIP_DEFLATED
                        ) as zip_file:

                            for index, task in enumerate(tasks):

                                try:

                                    status.info(
                                        f"İşleniyor: {index + 1}/{len(tasks)}"
                                    )

                                    response = session.get(
                                        task["url"],
                                        timeout=30,
                                        headers={
                                            "User-Agent":
                                            "Mozilla/5.0 Sistemist Image Studio"
                                        }
                                    )

                                    response.raise_for_status()

                                    image = Image.open(
                                        io.BytesIO(response.content)
                                    )

                                    image.load()

                                    processed_image = prepare_image(
                                        image,
                                        target_size,
                                        fit_mode_label
                                    )

                                    image_bytes, extension = save_image_to_buffer(
                                        processed_image,
                                        output_format,
                                        quality
                                    )

                                    filename = (
                                        f"{task['base']}"
                                        f"-{task['image_number']}"
                                        f"{extension}"
                                    )

                                    zip_file.writestr(
                                        filename,
                                        image_bytes
                                    )

                                    success_count += 1

                                except Exception as error:

                                    failed_count += 1

                                    errors.append({
                                        "URL": task["url"],
                                        "Hata": str(error)
                                    })

                                progress.progress(
                                    (index + 1) / len(tasks)
                                )

                        zip_buffer.seek(0)

                        status.empty()

                        if success_count > 0:

                            add_history(
                                "URL → Görsel",
                                "Başarılı",
                                f"{success_count} görsel indirildi ve işlendi",
                                success_count
                            )

                            st.success(
                                f"İşlem tamamlandı. {success_count} görsel başarıyla işlendi."
                            )

                            st.download_button(
                                "ZIP DOSYASINI İNDİR",
                                data=zip_buffer.getvalue(),
                                file_name=(
                                    "sistemist-url-gorsel-"
                                    f"{datetime.now().strftime('%Y%m%d-%H%M%S')}.zip"
                                ),
                                mime="application/zip",
                                key="download_url_zip"
                            )

                        if failed_count > 0:

                            st.warning(
                                f"{failed_count} görsel indirilemedi veya işlenemedi."
                            )

                            if errors:

                                error_workbook = Workbook()
                                error_sheet = error_workbook.active

                                error_sheet.title = "Hatalar"

                                error_sheet.append([
                                    "URL",
                                    "HATA"
                                ])

                                for item in errors:
                                    error_sheet.append([
                                        item["URL"],
                                        item["Hata"]
                                    ])

                                error_buffer = io.BytesIO()

                                error_workbook.save(error_buffer)

                                st.download_button(
                                    "HATA RAPORUNU İNDİR",
                                    data=error_buffer.getvalue(),
                                    file_name="sistemist-hata-raporu.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

        except Exception as error:

            st.error(
                f"Excel işleme hatası: {str(error)}"
            )

    app_footer()


# =========================================================
# IMAGE -> URL
# =========================================================

elif st.session_state.current_page == "Görsel → URL":

    page_header(
        "Görsel → <span>URL</span> Bulut Merkezi",
        "Görsellerinizi Cloudflare R2'ye yükleyin ve otomatik oluşturulan URL listesini Excel olarak indirin.",
        "SİSTEMİST CLOUD ENGINE"
    )

    if not r2_is_configured():

        st.warning(
            "Cloudflare R2 ayarları henüz tamamlanmadı. Önce Cloud R2 Ayarları sayfasını doldurun."
        )

        if st.button(
            "CLOUD R2 AYARLARINA GİT",
            key="go_r2_from_upload"
        ):
            go_to("Cloud R2 Ayarları")
            st.rerun()

    else:

        st.success(
            f"Cloud R2 bağlantısı yapılandırıldı. Bucket: {st.session_state.r2_bucket}"
        )

        upload_folder = st.text_input(
            "R2 klasörü / prefix (isteğe bağlı)",
            value="uploads",
            key="upload_folder"
        )

        uploaded_images = st.file_uploader(
            "Görselleri sürükleyin veya seçin",
            type=[
                "jpg", "jpeg", "png",
                "webp", "gif", "bmp"
            ],
            accept_multiple_files=True,
            key="r2_image_uploader"
        )

        if uploaded_images:

            st.info(
                f"{len(uploaded_images)} görsel yüklenmeye hazır."
            )

            if st.button(
                "BULUT YÜKLEMESİNİ BAŞLAT VE EXCEL RAPORU OLUŞTUR",
                key="upload_to_r2"
            ):

                try:

                    s3_client = get_r2_client()

                    results = []

                    success_count = 0
                    failed_count = 0

                    progress = st.progress(0)
                    status = st.empty()

                    for index, uploaded_file in enumerate(uploaded_images):

                        original_name = clean_filename(
                            Path(uploaded_file.name).stem
                        )

                        extension = Path(
                            uploaded_file.name
                        ).suffix.lower()

                        if not extension:
                            extension = ".jpg"

                        # Aynı isimli dosyaların birbirini ezmesini önlemek için
                        # her yüklemeye benzersiz bir zaman damgası eklenir.
                        unique_stamp = datetime.now().strftime("%H%M%S%f")
                        filename = (
                            f"{original_name}-{unique_stamp}"
                            f"{extension}"
                        )

                        timestamp_prefix = (
                            datetime.now()
                            .strftime("%Y/%m/%d")
                        )

                        clean_folder = (
                            upload_folder
                            .strip("/")
                            .strip()
                        )

                        if clean_folder:

                            object_key = (
                                f"{clean_folder}/"
                                f"{timestamp_prefix}/"
                                f"{filename}"
                            )

                        else:

                            object_key = (
                                f"{timestamp_prefix}/"
                                f"{filename}"
                            )

                        try:

                            status.info(
                                f"Yükleniyor: {index + 1}/{len(uploaded_images)}"
                            )

                            file_data = uploaded_file.getvalue()

                            content_type = (
                                mimetypes.guess_type(
                                    uploaded_file.name
                                )[0]
                                or "application/octet-stream"
                            )

                            s3_client.put_object(
                                Bucket=st.session_state.r2_bucket,
                                Key=object_key,
                                Body=file_data,
                                ContentType=content_type
                            )

                            public_url = build_public_url(
                                object_key
                            )

                            results.append([
                                uploaded_file.name,
                                object_key,
                                Path(uploaded_file.name)
                                .suffix
                                .replace(".", "")
                                .upper(),
                                round(
                                    len(file_data) / 1048576,
                                    3
                                ),
                                public_url,
                                "BAŞARILI"
                            ])

                            success_count += 1

                        except Exception as error:

                            results.append([
                                uploaded_file.name,
                                "",
                                Path(uploaded_file.name).suffix.replace(".", "").upper(),
                                "",
                                "",
                                f"HATA: {str(error)}"
                            ])

                            failed_count += 1

                        progress.progress(
                            (index + 1) / len(uploaded_images)
                        )

                    status.empty()

                    workbook = Workbook()
                    worksheet = workbook.active

                    worksheet.title = "Image URLs"

                    worksheet.append([
                        "DOSYA_ADI",
                        "R2_OBJECT_KEY",
                        "FORMAT",
                        "BOYUT_MB",
                        "URL",
                        "DURUM"
                    ])

                    for row in results:
                        worksheet.append(row)

                    worksheet.freeze_panes = "A2"

                    for column in worksheet.columns:

                        max_length = 0
                        column_letter = column[0].column_letter

                        for cell in column:

                            try:
                                max_length = max(
                                    max_length,
                                    len(str(cell.value))
                                )
                            except Exception:
                                pass

                        worksheet.column_dimensions[
                            column_letter
                        ].width = min(
                            max_length + 2,
                            80
                        )

                    excel_buffer = io.BytesIO()

                    workbook.save(excel_buffer)

                    excel_buffer.seek(0)

                    if success_count > 0:

                        add_history(
                            "Görsel → URL",
                            "Başarılı",
                            f"{success_count} görsel Cloudflare R2'ye yüklendi",
                            success_count
                        )

                        st.success(
                            f"Yükleme tamamlandı. {success_count} görsel başarıyla Cloudflare R2'ye gönderildi."
                        )

                        st.download_button(
                            "EXCEL URL RAPORUNU İNDİR",
                            data=excel_buffer.getvalue(),
                            file_name=(
                                "sistemist-r2-url-raporu-"
                                f"{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
                            ),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_r2_excel"
                        )

                    if failed_count:

                        st.warning(
                            f"{failed_count} dosya yüklenemedi. Excel raporunda hata bilgileri bulunmaktadır."
                        )

                except Exception as error:

                    st.error(
                        f"Cloudflare R2 bağlantı hatası: {str(error)}"
                    )

    app_footer()


# =========================================================
# BATCH CONVERSION
# =========================================================

elif st.session_state.current_page == "Toplu Dönüştürme":

    page_header(
        "<span>Toplu Görsel</span> Dönüştürme",
        "Bilgisayarınızdaki görselleri toplu olarak yeniden boyutlandırın, dönüştürün ve ZIP dosyası olarak indirin.",
        "SİSTEMİST BATCH ENGINE"
    )

    uploaded_images = st.file_uploader(
        "Görselleri seçin",
        type=[
            "jpg", "jpeg", "png",
            "webp", "gif", "bmp"
        ],
        accept_multiple_files=True,
        key="batch_images"
    )

    if uploaded_images:

        st.success(
            f"{len(uploaded_images)} görsel seçildi."
        )

        col1, col2, col3 = st.columns(3)

        with col1:

            batch_format = st.selectbox(
                "Yeni format",
                ["JPG", "PNG", "WEBP"],
                key="batch_format"
            )

        with col2:

            batch_size = st.selectbox(
                "Yeni boyut",
                [
                    "1200 × 1200 px",
                    "1200 × 1800 px",
                    "1000 × 1000 px",
                    "800 × 800 px",
                    "1920 × 1920 px",
                    "Orijinal Boyut"
                ],
                key="batch_size"
            )

        with col3:

            batch_quality = st.slider(
                "Kalite",
                60,
                100,
                90,
                key="batch_quality"
            )

        batch_fit = st.selectbox(
            "Yerleşim yöntemi",
            [
                "Sığdır",
                "Kırp"
            ],
            key="batch_fit"
        )

        if st.button(
            "TOPLU DÖNÜŞTÜRMEYİ BAŞLAT",
            key="start_batch"
        ):

            target_size = get_target_size(
                batch_size
            )

            zip_buffer = io.BytesIO()

            success_count = 0
            failed_count = 0

            progress = st.progress(0)

            with zipfile.ZipFile(
                zip_buffer,
                "w",
                zipfile.ZIP_DEFLATED
            ) as zip_file:

                for index, uploaded_file in enumerate(
                    uploaded_images
                ):

                    try:

                        image = Image.open(
                            io.BytesIO(
                                uploaded_file.getvalue()
                            )
                        )

                        image.load()

                        processed_image = prepare_image(
                            image,
                            target_size,
                            batch_fit
                        )

                        image_bytes, extension = save_image_to_buffer(
                            processed_image,
                            batch_format,
                            batch_quality
                        )

                        base_name = clean_filename(
                            Path(
                                uploaded_file.name
                            ).stem
                        )

                        output_name = (
                            f"{base_name}{extension}"
                        )

                        zip_file.writestr(
                            output_name,
                            image_bytes
                        )

                        success_count += 1

                    except Exception:
                        failed_count += 1

                    progress.progress(
                        (index + 1)
                        / len(uploaded_images)
                    )

            zip_buffer.seek(0)

            if success_count:

                add_history(
                    "Toplu Dönüştürme",
                    "Başarılı",
                    f"{success_count} görsel dönüştürüldü",
                    success_count
                )

                st.success(
                    f"{success_count} görsel başarıyla dönüştürüldü."
                )

                st.download_button(
                    "DÖNÜŞTÜRÜLEN GÖRSELLERİ İNDİR",
                    data=zip_buffer.getvalue(),
                    file_name=(
                        "sistemist-toplu-donusum-"
                        f"{datetime.now().strftime('%Y%m%d-%H%M%S')}.zip"
                    ),
                    mime="application/zip"
                )

            if failed_count:

                st.warning(
                    f"{failed_count} görsel işlenemedi."
                )

    app_footer()


# =========================================================
# HISTORY
# =========================================================

elif st.session_state.current_page == "İşlem Geçmişi":

    page_header(
        "İşlem <span>Geçmişi</span>",
        "Sistem üzerinde gerçekleştirdiğiniz görsel indirme, dönüştürme ve Cloudflare R2 yükleme operasyonlarını takip edin.",
        "OPERASYON KAYITLARI"
    )

    if not st.session_state.history:

        st.info(
            "Henüz kayıtlı bir işlem bulunmuyor."
        )

    else:

        col1, col2 = st.columns(
            [4, 1]
        )

        with col2:

            if st.button(
                "GEÇMİŞİ TEMİZLE",
                key="clear_history"
            ):

                st.session_state.history = []

                st.rerun()

        st.dataframe(
            st.session_state.history,
            use_container_width=True,
            hide_index=True
        )

    app_footer()


# =========================================================
# CLOUD FILES
# =========================================================

elif st.session_state.current_page == "Cloud Dosyaları":

    page_header(
        "<span>Cloud</span> Dosyaları",
        "Cloudflare R2 bucket içerisindeki dosyaları görüntüleyin, URL'lerini kopyalayın ve gerekli dosyaları silin.",
        "R2 STORAGE MANAGER"
    )

    if not r2_is_configured():

        st.warning(
            "Cloud R2 ayarları tamamlanmamış."
        )

        if st.button(
            "R2 AYARLARINA GİT",
            key="go_r2_cloud_files"
        ):
            go_to("Cloud R2 Ayarları")
            st.rerun()

    else:

        prefix = st.text_input(
            "Klasör / Prefix filtresi",
            value="",
            placeholder="Örn: uploads/2026"
        )

        if st.button(
            "CLOUD DOSYALARINI YÜKLE",
            key="load_cloud_files"
        ):

            try:

                client = get_r2_client()

                response = client.list_objects_v2(
                    Bucket=st.session_state.r2_bucket,
                    Prefix=prefix.strip()
                )

                contents = response.get(
                    "Contents",
                    []
                )

                if not contents:

                    st.info(
                        "Bu klasörde dosya bulunamadı."
                    )

                else:

                    st.success(
                        f"{len(contents)} dosya bulundu."
                    )

                    files = []

                    for item in contents:

                        key = item["Key"]

                        files.append({
                            "DOSYA": key,
                            "BOYUT": format_size(
                                item.get("Size", 0)
                            ),
                            "TARİH": item[
                                "LastModified"
                            ].strftime(
                                "%d.%m.%Y %H:%M"
                            ),
                            "URL": build_public_url(
                                key
                            )
                        })

                    st.dataframe(
                        files,
                        use_container_width=True,
                        hide_index=True
                    )

                    st.markdown("---")

                    st.subheader(
                        "Dosya Sil"
                    )

                    file_keys = [
                        item["DOSYA"]
                        for item in files
                    ]

                    selected_key = st.selectbox(
                        "Silinecek dosya",
                        file_keys,
                        key="delete_cloud_select"
                    )

                    if st.button(
                        "SEÇİLEN DOSYAYI SİL",
                        key="delete_cloud_file"
                    ):

                        try:

                            client.delete_object(
                                Bucket=st.session_state.r2_bucket,
                                Key=selected_key
                            )

                            add_history(
                                "Cloud Dosya Silme",
                                "Başarılı",
                                selected_key,
                                1
                            )

                            st.success(
                                "Dosya Cloudflare R2'den silindi."
                            )

                        except Exception as error:

                            st.error(
                                f"Silme hatası: {str(error)}"
                            )

            except Exception as error:

                st.error(
                    f"Cloud dosyaları alınamadı: {str(error)}"
                )

    app_footer()


# =========================================================
# R2 SETTINGS
# =========================================================

elif st.session_state.current_page == "Cloud R2 Ayarları":

    page_header(
        "Cloudflare <span>R2 Ayarları</span>",
        "Cloudflare R2 API bilgilerinizi girin, bağlantıyı test edin ve Sistemist Image Studio bulut depolamasını aktif hale getirin.",
        "CLOUD INFRASTRUCTURE"
    )

    st.markdown(
        dedent("""
        <div class="panel">
            <div class="panel-title">Cloudflare R2 API Bilgileri</div>
            <div class="panel-subtitle">
                Bu bilgiler Cloudflare hesabınızdan oluşturduğunuz R2 API Token
                ve bucket yapılandırmasına göre girilmelidir.
            </div>
        </div>
        """),
        unsafe_allow_html=True
    )

    endpoint = st.text_input(
        "R2 Endpoint",
        value=st.session_state.r2_endpoint,
        placeholder="https://ACCOUNT_ID.r2.cloudflarestorage.com"
    )

    access_key = st.text_input(
        "Access Key ID",
        value=st.session_state.r2_access_key
    )

    secret_key = st.text_input(
        "Secret Access Key",
        value=st.session_state.r2_secret_key,
        type="password"
    )

    bucket = st.text_input(
        "Bucket Name",
        value=st.session_state.r2_bucket
    )

    public_url = st.text_input(
        "CDN / Public URL",
        value=st.session_state.r2_public_url,
        placeholder="https://images.sistemist.com"
    )

    region = st.text_input(
        "Region",
        value=st.session_state.r2_region
    )

    col1, col2 = st.columns(2)

    with col1:

        if st.button(
            "AYARLARI KAYDET",
            key="save_r2_settings"
        ):

            st.session_state.r2_endpoint = endpoint.strip()
            st.session_state.r2_access_key = access_key.strip()
            st.session_state.r2_secret_key = secret_key.strip()
            st.session_state.r2_bucket = bucket.strip()
            st.session_state.r2_public_url = public_url.strip()
            st.session_state.r2_region = region.strip() or "auto"

            st.success(
                "Cloudflare R2 ayarları kaydedildi. Kalıcı kullanım için aynı bilgileri Render Environment Variables bölümüne de ekleyin."
            )

    with col2:

        if st.button(
            "BAĞLANTIYI TEST ET",
            key="test_r2_connection"
        ):

            st.session_state.r2_endpoint = endpoint.strip()
            st.session_state.r2_access_key = access_key.strip()
            st.session_state.r2_secret_key = secret_key.strip()
            st.session_state.r2_bucket = bucket.strip()
            st.session_state.r2_public_url = public_url.strip()
            st.session_state.r2_region = region.strip() or "auto"

            try:

                client = get_r2_client()

                client.head_bucket(
                    Bucket=st.session_state.r2_bucket
                )

                st.success(
                    "Cloudflare R2 bağlantısı başarılı. Bucket erişilebilir durumda."
                )

            except Exception as error:

                st.error(
                    f"Bağlantı kurulamadı: {str(error)}"
                )

    app_footer()


# =========================================================
# GENERAL SETTINGS
# =========================================================

elif st.session_state.current_page == "Genel Ayarlar":

    page_header(
        "<span>Genel</span> Ayarlar",
        "Sistemist Image Studio çalışma alanınızın genel ayarlarını yönetin.",
        "SİSTEM YAPILANDIRMASI"
    )

    st.markdown(
        dedent("""
        <div class="panel">
            <div class="panel-title">Uygulama Bilgileri</div>
            <div class="panel-subtitle">
                Sistemist Image Studio Web V7.7 PRO
            </div>
        </div>
        """),
        unsafe_allow_html=True
    )

    st.text_input(
        "Uygulama adı",
        value="Sistemist Image Studio Web"
    )

    st.selectbox(
        "Varsayılan çıktı formatı",
        ["JPG", "PNG", "WEBP"]
    )

    st.selectbox(
        "Varsayılan görsel boyutu",
        [
            "1200 × 1200 px",
            "1200 × 1800 px",
            "1000 × 1000 px"
        ]
    )

    st.success(
        "Uygulama şu anda profesyonel görsel operasyon modunda çalışıyor."
    )

    app_footer()


# =========================================================
# HELP CENTER
# =========================================================

elif st.session_state.current_page == "Yardım Merkezi":

    page_header(
        "Yardım <span>Merkezi</span>",
        "Sistemist Image Studio araçlarının nasıl kullanılacağını buradan takip edebilirsiniz.",
        "DESTEK"
    )

    with st.expander(
        "URL → Görsel nasıl kullanılır?",
        expanded=True
    ):

        st.write(
            """
            1. Excel dosyanızı yükleyin.
            2. Ürün adı veya stok kodu sütununu seçin.
            3. Görsel formatını seçin.
            4. Görsel ölçüsünü belirleyin.
            5. İşlemi başlatın.
            6. Oluşturulan ZIP dosyasını indirin.
            """
        )

    with st.expander(
        "Görsel → URL nasıl kullanılır?"
    ):

        st.write(
            """
            1. Önce Cloud R2 Ayarları bölümünden API bilgilerinizi girin.
            2. Görselleri seçin.
            3. Bulut yüklemesini başlatın.
            4. İşlem tamamlandığında Excel URL raporunu indirin.
            """
        )

    with st.expander(
        "Cloudflare R2 Endpoint nereden alınır?"
    ):

        st.write(
            """
            Cloudflare Dashboard → R2 Object Storage → Manage R2 API Tokens
            bölümünden Access Key ve Secret Key oluşturabilirsiniz.
            Endpoint değeri Cloudflare hesabınıza özel ACCOUNT_ID ile oluşturulur.
            """
        )

    app_footer()


# =========================================================
# PACKAGE
# =========================================================

elif st.session_state.current_page == "Paket & Lisans":

    page_header(
        "<span>Paket</span> & Lisans",
        "Sistemist Image Studio profesyonel SaaS altyapısı.",
        "ABONELİK YÖNETİMİ"
    )

    col1, col2, col3 = st.columns(3)

    packages = [
        {
            "name": "STARTER",
            "desc": "Temel görsel işlemleri",
            "price": "Başlangıç"
        },
        {
            "name": "PRO",
            "desc": "Tüm profesyonel araçlar",
            "price": "AKTİF"
        },
        {
            "name": "BUSINESS",
            "desc": "Yüksek hacimli operasyon",
            "price": "Kurumsal"
        }
    ]

    for column, package in zip(
        [col1, col2, col3],
        packages
    ):

        with column:

            st.markdown(
                dedent(f"""
                <div class="panel">
                    <div class="system-read">
                        {package["name"]}
                    </div>

                    <div class="panel-title">
                        {package["price"]}
                    </div>

                    <div class="panel-subtitle">
                        {package["desc"]}
                    </div>
                </div>
                """),
                unsafe_allow_html=True
            )

            if st.button(
                f"{package['name']} PAKETİNİ SEÇ",
                key=f"package_{package['name']}"
            ):

                st.session_state.active_package = (
                    package["name"]
                )

                st.success(
                    f"{package['name']} paketi aktif paket olarak seçildi."
                )

    app_footer()
